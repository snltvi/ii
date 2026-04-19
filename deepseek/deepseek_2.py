#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Расчет суточного пробега через API Mobiteam
Работает с Excel файлом, содержащим ID объектов и SID датчиков CAN
"""

import requests
import pandas as pd
from datetime import datetime, timedelta
import sys
import os
import json
import logging

# ============================================================================
# КОНСТАНТЫ
# ============================================================================

API_BASE_URL = "https://gps.mobiteam.com.ua/api/integration/v1"

# Credentials для автоматического получения SessionId
LOGIN = "abvprom"
PASSWORD = "29328"

# Поиск входного файла с датчиками
EXCEL_FILES = ['CAN_пробег_датчики_06_02_2026.xlsx', 'Датчики_CAN_пробег.xlsx']

# Настройка логгирования
def setup_logging():
    """Настройка логгирования"""
    logger = logging.getLogger(__name__)
    
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        
        # Формат для логов
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        
        # Файловый обработчик
        file_handler = logging.FileHandler(f'odometer_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log', encoding='utf-8')
        file_handler.setFormatter(formatter)
        
        # Консольный обработчик
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)
    
    return logger

# Инициализация логгера
logger = setup_logging()


# ============================================================================
# КЛАСС ДЛЯ КЭШИРОВАНИЯ ИНФОРМАЦИИ ОБ ОБЪЕКТАХ
# ============================================================================

class ObjectInfoCache:
    """Кэш информации об объектах для уменьшения количества запросов к API"""
    
    def __init__(self, session_id):
        self.session_id = session_id
        self.cache = {}
    
    def get_info(self, oid):
        """
        Получение информации об объекте из кэша или API
        
        Args:
            oid: ID объекта
            
        Returns:
            dict: {"name": ..., "group": ...} или None
        """
        if oid not in self.cache:
            self.cache[oid] = self._fetch_object_info(oid)
        return self.cache[oid]
    
    def _fetch_object_info(self, oid):
        """
        Получение информации об объекте из API
        
        Args:
            oid: ID объекта
            
        Returns:
            dict: {"name": ..., "group": ...} или None
        """
        url = f"{API_BASE_URL}/getobjectsstate"
        headers = {'SessionId': self.session_id}
        params = {'objuids': str(oid)}
        
        try:
            response = requests.get(url, headers=headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            if not data or len(data) == 0:
                logger.warning(f"Объект OID={oid} не найден в API")
                return None
            
            obj = data[0]
            return {
                'name': obj.get('name', f'Объект {oid}'),
                'group': obj.get('group', '')
            }
        except Exception as e:
            logger.error(f"Ошибка получения информации об объекте OID={oid}: {e}")
            return None


# ============================================================================
# ФУНКЦИИ API
# ============================================================================

def connect_to_api():
    """
    Подключение к API и получение SessionId
    
    Returns:
        str: SessionId для использования в заголовках запросов
    """
    url = f"{API_BASE_URL}/connect"
    params = {
        'login': LOGIN,
        'password': PASSWORD,
        'lang': 'ru-ru',
        'timezone': '3'
    }
    
    try:
        logger.info("Подключение к API Mobiteam...")
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        
        session_id = response.headers.get('sessionid')
        
        if not session_id:
            raise Exception("SessionId не найден в ответе API")
        
        logger.info(f"✓ SessionId получен: {session_id[:10]}...")
        return session_id
        
    except requests.exceptions.Timeout:
        logger.error("Таймаут подключения к API")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Ошибка подключения к API: {e}")
        sys.exit(1)


def get_odometer_data(session_id, oid, sensor_id, date_from, date_to):
    """
    Получение данных одометра напрямую через датчик CAN
    с улучшенной проверкой качества данных
    
    Args:
        session_id: SessionId для аутентификации
        oid: ID объекта
        sensor_id: ID датчика (SID)
        date_from: начало периода (строка 'YYYY-MM-DD HH:MM:SS')
        date_to: конец периода (строка 'YYYY-MM-DD HH:MM:SS')
    
    Returns:
        tuple: (odo_start, odo_end, mileage, records_count, time_span_hours)
    """
    url = f"{API_BASE_URL}/objdata"
    
    params = {
        'oid': oid,
        'slist': f's{sensor_id}',
        'from': date_from,
        'to': date_to,
        'compress': 'true'  # Сжатие данных для уменьшения объема
    }
    
    headers = {'SessionId': session_id}
    
    try:
        logger.debug(f"Запрос данных OID={oid}, SID={sensor_id}, период: {date_from} - {date_to}")
        response = requests.get(url, headers=headers, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        if data.get('result') != 'Ok':
            error_msg = data.get('error', 'Неизвестная ошибка')
            logger.warning(f"API вернуло ошибку для OID={oid}: {error_msg}")
            return None, None, None, 0, 0
        
        records = data.get('obj_data', {}).get('records', [])
        
        if not records:
            logger.debug(f"Нет записей одометра для OID={oid} за указанный период")
            return None, None, None, 0, 0
        
        # Фильтруем записи с непустыми значениями
        valid_records = []
        for rec in records:
            if len(rec) >= 2 and rec[1] is not None:
                try:
                    # Пробуем преобразовать в число
                    val = float(rec[1])
                    # Проверяем на разумные значения (не миллионы км)
                    if 0 <= val <= 9999999:
                        valid_records.append((rec[0], val))
                    else:
                        logger.warning(f"Аномальное значение одометра OID={oid}: {val}")
                except (ValueError, TypeError) as e:
                    logger.debug(f"Некорректное значение одометра OID={oid}: {rec[1]} - {e}")
                    continue
        
        records_count = len(valid_records)
        
        if records_count < 2:
            logger.debug(f"Мало валидных записей для OID={oid}: {records_count}")
            return None, None, None, records_count, 0
        
        # Сортируем по времени (на всякий случай)
        valid_records.sort(key=lambda x: x[0])
        
        value_start = float(valid_records[0][1])
        value_end = float(valid_records[-1][1])
        mileage = value_end - value_start
        
        # Расчет временного охвата данных
        try:
            # Преобразуем строки времени в datetime
            first_time_str = valid_records[0][0].replace('Z', '+00:00')
            last_time_str = valid_records[-1][0].replace('Z', '+00:00')
            
            first_time = datetime.fromisoformat(first_time_str)
            last_time = datetime.fromisoformat(last_time_str)
            
            time_span = last_time - first_time
            time_span_hours = time_span.total_seconds() / 3600
            
            # Логируем охват данных
            if time_span_hours < 20:
                logger.warning(f"Данные OID={oid} только за {time_span_hours:.1f} часов")
        except Exception as e:
            logger.debug(f"Ошибка расчета временного охвата OID={oid}: {e}")
            time_span_hours = 0
        
        # Дополнительные проверки
        if mileage < 0:
            logger.warning(f"Отрицательный пробег OID={oid}: {mileage:.2f} км")
            # Иногда одометр сбрасывается - считаем от 0
            if value_end >= 0:
                mileage = value_end
                logger.info(f"Корректировка пробега OID={oid} на {value_end:.2f} км")
        
        # Проверка на аномалии (более 2000 км/сутки)
        if mileage > 2000:
            logger.warning(f"Аномальный пробег OID={oid}: {mileage:.2f} км")
        
        logger.debug(f"Успешно получены данные OID={oid}: {mileage:.2f} км ({records_count} записей)")
        return value_start, value_end, mileage, records_count, time_span_hours
        
    except requests.exceptions.Timeout:
        logger.error(f"Таймаут запроса для OID={oid}")
        return None, None, None, 0, 0
    except Exception as e:
        logger.error(f"Ошибка получения данных OID={oid}: {str(e)[:100]}")
        return None, None, None, 0, 0


# ============================================================================
# РАСЧЕТ ПРОБЕГА - РЕЖИМ С EXCEL
# ============================================================================

def calculate_from_excel(session_id, vehicles_df, target_date):
    """
    Расчет пробега из списка в Excel файле (требует колонку SID)
    
    Args:
        session_id: SessionId для аутентификации
        vehicles_df: DataFrame с данными об автомобилях (должна содержать 'ID объекта' и 'SID')
        target_date: целевая дата (datetime)
    
    Returns:
        DataFrame с результатами
    """
    date_str = target_date.strftime('%Y-%m-%d')
    date_from = f"{date_str} 00:00:00"
    date_to = f"{date_str} 23:59:59"  # 59 секунд для полного дня
    
    logger.info(f"Начало расчета пробега за {date_str}")
    logger.info(f"Период UTC: {date_from} - {date_to}")
    
    print(f"\n{'='*80}")
    print(f"РАСЧЕТ ПРОБЕГА ЗА {target_date.strftime('%d.%m.%Y')}")
    print(f"{'='*80}\n")
    
    results = []
    cache = ObjectInfoCache(session_id)
    total_vehicles = len(vehicles_df)
    
    # Простой прогресс-бар без tqdm
    for idx, row in vehicles_df.iterrows():
        try:
            # Показываем прогресс
            progress = (idx + 1) / total_vehicles * 100
            print(f"\rОбработка ТС: {idx + 1}/{total_vehicles} ({progress:.1f}%)", end='', flush=True)
            
            oid = int(row['ID объекта'])
            sensor_id = int(row['SID'])
            driver = row.get('ФИО', '')
            vehicle_number = row.get('Номер авто', '')
            trailer_number = row.get('Номер прицепа', '')
            
            # Получаем название объекта, если не указано
            if not driver and not vehicle_number:
                obj_info = cache.get_info(oid)
                if obj_info:
                    vehicle_number = obj_info['name']
                    driver = obj_info.get('group', '')
                else:
                    vehicle_number = f'ID_{oid}'
            
            # Получение данных одометра
            odo_start, odo_end, mileage, records_count, time_span_hours = get_odometer_data(
                session_id, oid, sensor_id, date_from, date_to
            )
            
            # Определяем статус
            if odo_start is None or records_count < 2:
                status = 'Нет данных'
                mileage_val = 0.0
                color_code = 'error'
            elif mileage > 2000 or mileage < 0:
                status = 'Проверить'
                mileage_val = round(mileage, 2) if mileage else 0.0
                color_code = 'warning'
            elif time_span_hours < 20:
                status = 'Неполный день'
                mileage_val = round(mileage, 2) if mileage else 0.0
                color_code = 'warning'
            else:
                status = 'OK'
                mileage_val = round(mileage, 2) if mileage else 0.0
                color_code = 'ok'
            
            results.append({
                '№': idx + 1,
                'ФИО': driver,
                'Номер авто': vehicle_number,
                'Номер прицепа': trailer_number,
                'ID объекта': oid,
                'SID': sensor_id,
                'Одометр начало (км)': round(odo_start, 2) if odo_start else None,
                'Одометр конец (км)': round(odo_end, 2) if odo_end else None,
                'Пробег (км)': mileage_val,
                'Записей': records_count,
                'Охват (часы)': round(time_span_hours, 1) if time_span_hours else 0,
                'Статус': status,
                'Цвет статуса': color_code,
                'Дата': date_str
            })
            
        except Exception as e:
            logger.error(f"Критическая ошибка при обработке строки {idx}: {e}")
            print(f"\nОшибка при обработке строки {idx}: {e}")
            
            results.append({
                '№': idx + 1,
                'ФИО': str(row.get('ФИО', '')),
                'Номер авто': str(row.get('Номер авто', '')),
                'Номер прицепа': str(row.get('Номер прицепа', '')),
                'ID объекта': int(row['ID объекта']) if 'ID объекта' in row else 0,
                'SID': int(row['SID']) if 'SID' in row else 0,
                'Одометр начало (км)': None,
                'Одометр конец (км)': None,
                'Пробег (км)': 0.0,
                'Записей': 0,
                'Охват (часы)': 0,
                'Статус': f'Ошибка: {str(e)[:30]}',
                'Цвет статуса': 'error',
                'Дата': date_str
            })
    
    print()  # Новая строка после прогресс-бара
    logger.info(f"Завершение расчета, обработано {len(results)} записей")
    return pd.DataFrame(results)


# ============================================================================
# СОХРАНЕНИЕ РЕЗУЛЬТАТОВ С ФОРМАТИРОВАНИЕМ
# ============================================================================

def save_results_with_format(results_df, output_file):
    """
    Сохранение результатов с цветовым форматированием
    
    Args:
        results_df: DataFrame с результатами
        output_file: имя выходного файла
    """
    try:
        # Удаляем служебную колонку перед сохранением
        if 'Цвет статуса' in results_df.columns:
            results_df_for_save = results_df.drop(columns=['Цвет статуса'])
        else:
            results_df_for_save = results_df
        
        # Создаем writer для Excel
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        results_df_for_save.to_excel(writer, index=False, sheet_name='Пробег')
        
        workbook = writer.book
        worksheet = writer.sheets['Пробег']
        
        # Автоширина колонок (упрощенный вариант без xlsxwriter)
        from openpyxl.utils import get_column_letter
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Простое цветовое форматирование
        from openpyxl.styles import PatternFill, Font
        
        # Определяем стили
        fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fill_warning = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        fill_error = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        font_ok = Font(color='006100')
        font_warning = Font(color='9C6500')
        font_error = Font(color='9C0006')
        
        # Находим колонку "Статус"
        status_col_idx = None
        for idx, col in enumerate(results_df_for_save.columns):
            if col == 'Статус':
                status_col_idx = idx + 1  # +1 потому что Excel начинается с 1
                break
        
        # Применяем форматирование к статусам
        if status_col_idx and 'Цвет статуса' in results_df.columns:
            for row_idx, color_code in enumerate(results_df['Цвет статуса'], start=2):  # start=2 потому что заголовок в 1 строке
                cell = worksheet.cell(row=row_idx, column=status_col_idx)
                
                if color_code == 'ok':
                    cell.fill = fill_ok
                    cell.font = font_ok
                elif color_code == 'warning':
                    cell.fill = fill_warning
                    cell.font = font_warning
                elif color_code == 'error':
                    cell.fill = fill_error
                    cell.font = font_error
        
        # Добавляем лист с итогами
        if len(results_df) > 0:
            summary_data = {
                'Метрика': [
                    'Всего ТС',
                    'Успешно (OK)',
                    'Требуют проверки',
                    'Нет данных',
                    'С ошибкой',
                    'Общий пробег (км)',
                    'Средний пробег (км)',
                    'Максимальный пробег (км)',
                    'Минимальный пробег (км)'
                ],
                'Значение': [
                    len(results_df),
                    len(results_df[results_df['Статус'] == 'OK']),
                    len(results_df[results_df['Статус'].isin(['Проверить', 'Неполный день'])]),
                    len(results_df[results_df['Статус'] == 'Нет данных']),
                    len(results_df[results_df['Статус'].str.contains('Ошибка', na=False)]),
                    results_df['Пробег (км)'].sum(),
                    results_df['Пробег (км)'].mean(),
                    results_df['Пробег (км)'].max(),
                    results_df['Пробег (км)'].min()
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name='Итоги')
        
        writer.close()
        logger.info(f"Результаты сохранены в файл: {output_file}")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка при сохранении результатов: {e}")
        # Пробуем сохранить без форматирования
        try:
            results_df.to_excel(output_file, index=False)
            logger.info(f"Результаты сохранены без форматирования в: {output_file}")
            return True
        except Exception as e2:
            logger.error(f"Не удалось сохранить файл: {e2}")
            return False


# ============================================================================
# ОБРАБОТКА ДИАПАЗОНА ДАТ
# ============================================================================

def process_date_range(session_id, vehicles_df, start_date, end_date):
    """
    Обработка диапазона дат
    
    Args:
        session_id: SessionId для аутентификации
        vehicles_df: DataFrame с данными об автомобилях
        start_date: начальная дата (datetime)
        end_date: конечная дата (datetime)
    
    Returns:
        tuple: (combined_df, summary_df)
    """
    logger.info(f"Начало обработки диапазона дат: {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}")
    
    current_date = start_date
    all_results = []
    
    while current_date <= end_date:
        logger.info(f"Обработка {current_date.strftime('%Y-%m-%d')}")
        print(f"\nОбработка даты: {current_date.strftime('%d.%m.%Y')}")
        
        try:
            daily_df = calculate_from_excel(session_id, vehicles_df, current_date)
            daily_df['Дата'] = current_date.strftime('%Y-%m-%d')
            all_results.append(daily_df)
            
            # Краткая статистика за день
            if len(daily_df) > 0:
                total_mileage = daily_df['Пробег (км)'].sum()
                ok_count = len(daily_df[daily_df['Статус'] == 'OK'])
                print(f"  Результат: {ok_count}/{len(daily_df)} OK, пробег: {total_mileage:.1f} км")
            
        except Exception as e:
            logger.error(f"Ошибка при обработке дня {current_date.strftime('%Y-%m-%d')}: {e}")
            print(f"  Ошибка: {e}")
        
        current_date += timedelta(days=1)
    
    if not all_results:
        logger.warning("Нет результатов за указанный период")
        return pd.DataFrame(), pd.DataFrame()
    
    # Объединяем все результаты
    combined_df = pd.concat(all_results, ignore_index=True)
    
    # Группировка по ТС за период
    try:
        summary_df = combined_df.groupby(['ФИО', 'Номер авто', 'ID объекта']).agg({
            'Пробег (км)': 'sum',
            'Статус': lambda x: 'OK' if all(s == 'OK' for s in x) else 'Есть проблемы',
            'Дата': 'count'
        }).reset_index()
        
        summary_df = summary_df.rename(columns={'Дата': 'Дней с данными'})
        
        logger.info(f"Обработка завершена. Всего дней: {len(all_results)}, записей: {len(combined_df)}")
        
        return combined_df, summary_df
        
    except Exception as e:
        logger.error(f"Ошибка при группировке данных: {e}")
        return combined_df, pd.DataFrame()


# ============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================================

def main():
    """Главная функция скрипта"""
    print("\n" + "="*80)
    print("РАСЧЁТ СУТОЧНОГО ПРОБЕГА - Mobiteam API v2.0")
    print("="*80)
    
    # Проверка наличия Excel файла
    excel_file = None
    for filename in EXCEL_FILES:
        if os.path.exists(filename):
            excel_file = filename
            logger.info(f"Найден файл: {filename}")
            break
    
    if not excel_file:
        print("\n✗ Excel файл не найден!")
        print("\n💡 Для работы скрипта необходим Excel файл с колонки:")
        print("   - ID объекта (обязательно)")
        print("   - SID (обязательно - ID датчика CAN)")
        print("   - ФИО (опционально)")
        print("   - Номер авто (опционально)")
        print("   - Номер прицепа (опционально)")
        print(f"\nОжидаемые имена файлов: {', '.join(EXCEL_FILES)}")
        logger.error("Excel файл не найден")
        sys.exit(1)
    
    # Ввод даты
    print("\n" + "-"*80)
    target_date_str = input("Введите дату (YYYY-MM-DD) или нажмите Enter для вчерашней даты: ").strip()
    
    if not target_date_str:
        target_date = datetime.now() - timedelta(days=1)
        print(f"Используем вчерашнюю дату: {target_date.strftime('%Y-%m-%d')}")
    else:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        except ValueError:
            print("✗ Ошибка: неверный формат даты. Используйте YYYY-MM-DD")
            logger.error(f"Неверный формат даты: {target_date_str}")
            sys.exit(1)
    
    # Опция обработки диапазона дат
    print("\n" + "-"*80)
    process_range = input("Обработать диапазон дат? (y/N): ").strip().lower()
    
    if process_range == 'y':
        try:
            start_date_str = input("Начальная дата (YYYY-MM-DD): ").strip()
            end_date_str = input("Конечная дата (YYYY-MM-DD): ").strip()
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            
            if start_date > end_date:
                print("✗ Ошибка: начальная дата позже конечной")
                logger.error(f"Некорректный диапазон дат: {start_date_str} > {end_date_str}")
                sys.exit(1)
                
            logger.info(f"Режим обработки диапазона: {start_date_str} - {end_date_str}")
            
        except ValueError:
            print("✗ Ошибка: неверный формат даты")
            logger.error(f"Ошибка ввода диапазона дат")
            sys.exit(1)
    
    # Подключение к API
    print("\n" + "-"*80)
    print("Подключение к API...")
    session_id = connect_to_api()
    
    # Загрузка Excel файла
    try:
        vehicles_df = pd.read_excel(excel_file)
        logger.info(f"Загружен файл: {excel_file}, записей: {len(vehicles_df)}")
        
        required = ['ID объекта', 'SID']
        missing = [col for col in required if col not in vehicles_df.columns]
        
        if missing:
            print(f"\n✗ ОШИБКА: Отсутствуют обязательные колонки: {missing}")
            print(f"  Найденные колонки: {list(vehicles_df.columns)}")
            logger.error(f"Отсутствуют обязательные колонки: {missing}")
            sys.exit(1)
        
        print(f"✓ Загружено {len(vehicles_df)} записей из {excel_file}")
        
    except Exception as e:
        print(f"✗ Ошибка чтения файла: {e}")
        logger.error(f"Ошибка чтения Excel файла: {e}")
        sys.exit(1)
    
    # Обработка данных
    if process_range == 'y':
        # Обработка диапазона дат
        combined_df, summary_df = process_date_range(session_id, vehicles_df, start_date, end_date)
        
        if combined_df.empty:
            print("\n✗ Нет данных за указанный период")
            logger.warning("Нет данных за указанный период")
            sys.exit(1)
        
        # Сохранение результатов
        output_file = f"Пробег_{start_date.strftime('%Y-%m-%d')}_по_{end_date.strftime('%Y-%m-%d')}.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            combined_df.to_excel(writer, sheet_name='Детализация', index=False)
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='Сводка', index=False)
            
            # Дополнительная статистика
            stats_df = pd.DataFrame({
                'Метрика': ['Начальная дата', 'Конечная дата', 'Всего дней', 'Всего ТС', 'Общий пробег (км)', 'Средний пробег в день (км)'],
                'Значение': [
                    start_date.strftime('%Y-%m-%d'),
                    end_date.strftime('%Y-%m-%d'),
                    (end_date - start_date).days + 1,
                    len(vehicles_df),
                    combined_df['Пробег (км)'].sum(),
                    combined_df['Пробег (км)'].sum() / ((end_date - start_date).days + 1)
                ]
            })
            stats_df.to_excel(writer, sheet_name='Статистика', index=False)
        
        print(f"\n✓ Результаты сохранены в {output_file}")
        
        # Вывод сводки
        print("\n" + "="*80)
        print("СВОДКА ПО ПЕРИОДУ")
        print("="*80)
        print(f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
        print(f"ТС в обработке: {len(vehicles_df)}")
        print(f"Дней обработано: {(end_date - start_date).days + 1}")
        print(f"Всего пробег: {combined_df['Пробег (км)'].sum():.1f} км")
        print(f"Средний пробег в день: {combined_df['Пробег (км)'].sum() / ((end_date - start_date).days + 1):.1f} км")
        
    else:
        # Обработка одной даты
        results_df = calculate_from_excel(session_id, vehicles_df, target_date)
        output_file = f"Пробег_{target_date.strftime('%Y-%m-%d')}.xlsx"
        
        # Сохранение с форматированием
        success = save_results_with_format(results_df, output_file)
        
        if not success:
            print("✗ Ошибка при сохранении результатов")
            logger.error("Не удалось сохранить результаты")
            sys.exit(1)
        
        # Вывод статистики
        print("\n" + "="*80)
        print("СТАТИСТИКА")
        print("="*80)
        
        if len(results_df) > 0:
            ok_count = len(results_df[results_df['Статус'] == 'OK'])
            check_count = len(results_df[results_df['Статус'].isin(['Проверить', 'Неполный день'])])
            no_data_count = len(results_df[results_df['Статус'] == 'Нет данных'])
            error_count = len(results_df[results_df['Статус'].str.contains('Ошибка', na=False)])
            
            print(f"Всего ТС: {len(results_df)}")
            print(f"✓ Успешно (OK): {ok_count}")
            print(f"⚠ Требуют проверки: {check_count}")
            print(f"✗ Нет данных: {no_data_count}")
            print(f"💥 С ошибкой: {error_count}")
            print(f"\n📊 ПРОБЕГ:")
            print(f"  Общий: {results_df['Пробег (км)'].sum():.2f} км")
            print(f"  Средний: {results_df['Пробег (км)'].mean():.2f} км")
            print(f"  Максимальный: {results_df['Пробег (км)'].max():.2f} км")
            print(f"  Минимальный: {results_df['Пробег (км)'].min():.2f} км")
            
            # Топ-5 по пробегу
            top5 = results_df.nlargest(5, 'Пробег (км)')
            print(f"\n🏆 ТОП-5 по пробегу:")
            for i, (_, row) in enumerate(top5.iterrows(), 1):
                print(f"  {i}. {row['Номер авто']}: {row['Пробег (км)']:.1f} км")
        
        print(f"\n✓ Результаты сохранены в файл: {output_file}")
    
    print("\n" + "="*80)
    print("ЗАВЕРШЕНО")
    print("="*80)
    logger.info("Скрипт успешно завершен")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n✗ Прервано пользователем")
        logger.info("Скрипт прерван пользователем")
        sys.exit(0)
    except Exception as e:
        print(f"\n✗ Непредвиденная ошибка: {e}")
        logger.exception(f"Непредвиденная ошибка: {e}")
        sys.exit(1)